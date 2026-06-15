#!/usr/bin/env python3
"""PR Monitor HTML + CSV 생성.

두 경로에서 자사 언급 기사 수집:
  A) Google News RSS 직접 fetch — 자사명·별칭 쿼리 (도메인팩 pr-queries 에서 구성)
  B) classify.py pr-articles 보완 (body 추출 성공한 건만)

주가/시황 기사 → 하단 별도 섹션.
"""

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import time
import urllib.parse
from datetime import date, datetime, timezone, timedelta
from pathlib import Path


def _resolve_claude_bin() -> str | None:
    """claude CLI 경로를 이식성 있게 해석.
    우선순위: 환경변수 CLAUDE_BIN → PATH(shutil.which) → 흔한 설치 경로.
    없으면 None → LLM 기능은 규칙 기반으로 graceful fallback.
    """
    env = os.environ.get("CLAUDE_BIN")
    if env and Path(env).exists():
        return env
    found = shutil.which("claude")
    if found:
        return found
    for cand in ("~/.local/bin/claude", "~/.claude/local/claude", "/usr/local/bin/claude", "/opt/homebrew/bin/claude"):
        p = Path(cand).expanduser()
        if p.exists():
            return str(p)
    return None


CLAUDE_BIN = _resolve_claude_bin()

# ── 도메인팩 로드 (엔진은 하드코딩된 org 지식을 갖지 않는다) ──────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from prmonitor import domainpack

_PR_QUERIES = domainpack.load_pack("pr-queries")
_TONE_LEXICON = domainpack.load_pack("tone-lexicon")
_MEDIA = domainpack.load_pack("media")
_BRANDING = domainpack.load_pack("branding")

# 기자명으로 오인되기 쉬운 매체·통신사명 (byline 추출 시 제외)
MEDIA_NAMES = set(_MEDIA["outlet_names"])


def _author_from_raw_html(html: str) -> str:
    """trafilatura가 본문에서 떼버린 byline을 raw HTML 메타/JSON-LD/리드에서 복구.

    우선순위: meta[name=author]·article:author → JSON-LD author.name → 리드 '이름 기자'.
    한글 이름이면 '이름 기자'로 정규화, 없으면 빈 문자열.
    """
    if not html:
        return ""

    def _to_kija(name: str) -> str:
        name = (name or "").strip()
        if not name:
            return ""
        m = re.search(r"([가-힣]{2,4})\s*기자(?![가-힣])", name)
        if m:
            cand = m.group(1)
            return "" if cand in MEDIA_NAMES else cand + " 기자"
        if re.fullmatch(r"[가-힣]{2,4}", name):  # 순수 한글 이름
            return "" if name in MEDIA_NAMES else name + " 기자"
        return ""  # 매체명·로마자 등은 버림

    # 1) 메타 태그
    for pat in (
        r'<meta[^>]+(?:name|property)=["\'](?:author|article:author|og:article:author|dable:author)["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+(?:name|property)=["\'](?:author|article:author)["\']',
    ):
        for mm in re.finditer(pat, html, re.IGNORECASE):
            got = _to_kija(mm.group(1))
            if got:
                return got

    # 2) JSON-LD "author":{"name":"홍길동 기자"} 또는 "author":"홍길동"
    for mm in re.finditer(r'"author"\s*:\s*\{[^}]*?"name"\s*:\s*"([^"]+)"', html):
        got = _to_kija(mm.group(1))
        if got:
            return got
    for mm in re.finditer(r'"author"\s*:\s*"([^"]+)"', html):
        got = _to_kija(mm.group(1))
        if got:
            return got

    # 3) 리드 byline: '[매체 이름 기자]' '(서울=뉴스1) 홍길동 기자'
    m = re.search(r"([가-힣]{2,4})\s*(?:기자|특파원)(?![가-힣])", html[:3000])
    if m:
        return m.group(1) + " 기자"
    return ""

# lib/ (gnews_resolver 등) 접근을 위해 scripts/ 디렉토리를 path에 추가
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    import feedparser
except ImportError:
    feedparser = None

from lib.common import CACHE_DIR, PR_OUTPUT_DIR, PROCESSED_DIR, load_json, save_json
PR_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
TODAY = sys.argv[1] if len(sys.argv) > 1 else str(date.today())
HOURS = int(sys.argv[2]) if len(sys.argv) > 2 else 48

# ── 키워드 설정 ─────────────────────────────────────────────────
SELF_KW = list(_PR_QUERIES["self_aliases"])

NEG_KW = list(_TONE_LEXICON["negative"])
POS_KW = list(_TONE_LEXICON["positive"])
# 제품·사업 진전 (출시/공급/납품/수상/선보 등은 명확한 긍정)
_POS_EXTRA = [
    "출시", "출고", "론칭", "공급", "납품", "선보", "공개", "출전",
    "수주", "양산", "출하", "런칭",
]
POS_KW = POS_KW + [kw for kw in _POS_EXTRA if kw not in POS_KW]
STOCK_KW = list(_TONE_LEXICON["stock"]) + ["시가총액", "종목"]

# Google News RSS — 자사 직접 쿼리 목록
GNEWS_QUERIES = [
    {"query": q["query"], "hl": q["hl"], "gl": q["gl"], "domestic": q["domestic"]}
    for q in _PR_QUERIES["gnews_queries"]
]

# 도메인 → 한국어 매체명 매핑
DOMAIN_TO_NAME: dict[str, str] = dict(_MEDIA["domain_to_name"])

GNEWS_TEMPLATE = "https://news.google.com/rss/search?q={query}&hl={hl}&gl={gl}&ceid={gl}:{hl}"


# ── 헬퍼 ────────────────────────────────────────────────────────

def is_within_hours(date_str: str, hours: int) -> bool:
    """published_date 가 hours 이내인지 확인."""
    if not date_str:
        return True  # 날짜 없으면 포함
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(hours=hours)
    try:
        for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%d"):
            try:
                if fmt == "%Y-%m-%d":
                    dt = datetime.strptime(date_str[:10], fmt).replace(tzinfo=timezone.utc)
                else:
                    dt = datetime.fromisoformat(date_str)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                return dt >= cutoff
            except ValueError:
                continue
    except Exception:
        pass
    return True


def detect_tone(title: str, body: str = "") -> str:
    """제목 기준 우선. 부정은 제목에 명확히 나올 때만, 긍정은 제목+본문."""
    t_lower = title.lower()
    # 부정: 제목에서만 판단 (본문 과민 방지)
    if any(k in t_lower for k in NEG_KW):
        return "부정"
    # 긍정: 제목 우선, 제목에 없으면 본문 앞 300자
    combined = (title + " " + body[:300]).lower()
    if any(k in combined for k in POS_KW):
        return "긍정"
    return "중립"


def classify_and_summarize_batch(rows: list[dict], n_direct: int) -> tuple[dict[int, str], dict[int, str]]:
    """haiku 1회 호출로 톤 판정(전체) + 직접 언급 1줄 요약(앞 n_direct건) 동시 수행.

    rows 의 0..n_direct-1 은 직접 언급 기사 (요약 대상).
    반환: ({idx: 긍정|중립|부정}, {idx: 요약}). 실패 시 빈 dict → 규칙 fallback.
    (과거 톤·요약을 별도 haiku 2회 호출 → CLI 콜드스타트·시스템프롬프트 중복 제거)
    """
    if not rows or not CLAUDE_BIN:
        return {}, {}
    org_name = _BRANDING["org_name"]
    lines = [
        f"각 기사가 '{org_name}' 입장에서 긍정·중립·부정 중 무엇인지 판정해.",
        f"추가로 0~{n_direct - 1}번 기사(자사 직접 언급)는 PR 관점 한 줄 요약도 붙여."
        if n_direct > 0 else "",
        "",
        "형식 (설명·머리말 없이 목록만, 번호는 0부터):",
        f"- 0~{n_direct - 1}번: '번호: 톤 | 요약'" if n_direct > 0 else "",
        f"- {n_direct}번 이후: '번호: 톤'" if n_direct > 0 else "- '번호: 톤'",
        "",
        "톤 판정 기준 (자사 관점):",
        "- 긍정: 자사 호재(수주·협력·투자유치·수상·기술인정·주가강세·우호적 평가·유망주 거론).",
        "- 부정: 자사 악재(검찰·소송·결함·리콜·사업차질·지연·실적부진·주가급락·부정적 평가·우려).",
        "- 중립: 단순 사실 보도·시황 나열·여러 기업 중 하나로 곁다리 언급·판단 어려움.",
        "주의: 제목의 전반적 뉘앙스를 우선하라. 본문 일부 단어에 휘둘리지 마라.",
        "      예) '숨고르기 들어간 로봇주' '긴장 고조' '먹구름'은 긍정이 아니다(중립/부정).",
        "요약 규칙: 45자 이내. 사실만. 제목 반복 금지, 본문 핵심(수치·상대·맥락). 추측·평가·전망 금지. 한국어.",
        "---",
    ]
    for i, r in enumerate(rows):
        if i < n_direct:
            extra = (r.get("_body_head", "") or r.get("summary", "") or "")[:300]
            lines.append(f"{i}: {r.get('title','')}" + (f" | 본문: {extra}" if extra else ""))
        else:
            ev = (r.get("evidence", "") or "")[:100]
            lines.append(f"{i}: {r.get('title','')}" + (f" | {ev}" if ev else ""))
    try:
        res = subprocess.run(
            [CLAUDE_BIN, "-p", "\n".join(lines), "--model", "haiku"],
            capture_output=True, text=True, timeout=180,
        )
        if res.returncode != 0:
            print(f"  WARN: 톤·요약 판정 실패 (rc={res.returncode}) → 규칙 fallback", file=sys.stderr)
            return {}, {}
        tones: dict[int, str] = {}
        summaries: dict[int, str] = {}
        for line in res.stdout.strip().splitlines():
            m = re.match(r"^(\d+)\s*[:：]\s*(긍정|중립|부정)(?:\s*\|\s*(.+))?", line.strip())
            if m:
                idx = int(m.group(1))
                tones[idx] = m.group(2)
                if m.group(3):
                    summaries[idx] = m.group(3).strip()
        print(f"  톤 판정: {len(tones)}/{len(rows)}건, 직접 요약: {len(summaries)}/{n_direct}건 (haiku 1회)",
              file=sys.stderr)
        return tones, summaries
    except Exception as e:
        print(f"  WARN: 톤·요약 판정 예외: {e} → 규칙 fallback", file=sys.stderr)
        return {}, {}


def is_stock_article(title: str, body: str = "") -> bool:
    """제목에 STOCK_KW 있으면 즉시 주가 기사.
    본문은 2개 이상 keyword 일치할 때만 주가 기사로 분류 (본문 과민 반응 방지)."""
    title_lower = title.lower()
    if any(k in title_lower for k in STOCK_KW):
        return True
    if body:
        body_lower = body.lower()
        hits = sum(1 for k in STOCK_KW if k in body_lower)
        return hits >= 2
    return False


def normalize_author(author: str) -> str:
    """기자명 통일/정제: '이름 기자' 형식.

    - 'Newsis; 뉴시스; 김경택 기자' → '김경택 기자' (한글이름+기자 우선)
    - '김태환 기자 Kth Kr 기자' → '김태환 기자' (첫 한글이름+기자)
    - 로마자·매체명 중복/세미콜론 나열 제거
    """
    if not author:
        return ""
    author = author.strip()
    if not author:
        return ""
    # 1) 본문 어디든 '한글이름(2~4자) + 기자' 패턴이 있으면 그것만 추출
    #    (기자 뒤 한글 차단: 기자단·기자회견 등 오매칭 방지)
    m = re.search(r"([가-힣]{2,4})\s*기자(?![가-힣])", author)
    if m and m.group(1) not in MEDIA_NAMES:
        return m.group(1) + " 기자"
    # 2) 세미콜론/슬래시 나열 → 한글 이름만 있는 토큰 선택
    parts = [p.strip() for p in re.split(r"[;/]", author) if p.strip()]
    for p in parts:
        if re.fullmatch(r"[가-힣]{2,4}", p):
            return p + " 기자"
    # 3) 한글 포함 토큰 → ' 기자' 부여
    for p in parts:
        if re.search(r"[가-힣]", p):
            return p if p.endswith("기자") else p + " 기자"
    # 4) 로마자/봇 등 → 빈 문자열 (garbage 노출 방지)
    return ""


def extract_evidence(title: str, body: str) -> str:
    """본문에서 자사 키워드 포함 문장 추출. 여러 곳이면 ' … '로 연결(최대 3개).

    - 소수점(.44%) 보호 위해 마침표+공백/줄바꿈 기준 분리.
    - 숫자·구두점·소문자로 시작하는 중간 파편은 제외.
    - 제목과 거의 동일한 문장 제외.
    """
    if not body:
        return ""
    sentences = [s.strip() for s in re.split(r"(?<=[.!?。])\s+|\n", body) if len(s.strip()) > 15]
    title_head = title[:80].strip()
    found: list[str] = []
    for s in sentences:
        if not any(kw in s.lower() for kw in SELF_KW):
            continue
        if s[:80].strip() == title_head:
            continue
        # fragment 필터: 한글·대문자·따옴표로 시작해야 정상 문장
        if not re.match(r'[A-Z가-힣“"\'‘]', s[0]):
            continue
        clip = s[:200].strip()
        if clip not in found:
            found.append(clip)
        if len(found) >= 3:
            break
    return " … ".join(found)


def extract_summary(body: str, evidence: str) -> str:
    """evidence 와 다른 첫 유효 문장."""
    sentences = [s.strip() for s in re.split(r"[.\n]", body) if len(s.strip()) > 20]
    for s in sentences:
        if s[:80] not in evidence[:80]:
            return s[:150]
    return ""


# ── Google News RSS 직접 수집 ────────────────────────────────────

def fetch_gnews_pr(hours: int) -> list[dict]:
    """Google News RSS에서 자사 언급 기사 직접 수집 (도메인팩 자사명·별칭 쿼리)."""
    if feedparser is None:
        print("  WARN: feedparser 없음 — pip install feedparser", file=sys.stderr)
        return []

    days = max(1, round(hours / 24))
    results: list[dict] = []
    seen_urls: set[str] = set()

    for q_cfg in GNEWS_QUERIES:
        query_with_window = f"{q_cfg['query']} when:{days}d"
        encoded = urllib.parse.quote_plus(query_with_window)
        url = GNEWS_TEMPLATE.format(
            query=encoded, hl=q_cfg["hl"], gl=q_cfg["gl"]
        )
        try:
            feed = feedparser.parse(url)
            entries = feed.entries or []
            print(f"  Google News [{q_cfg['query']}] → {len(entries)}건", file=sys.stderr)
        except Exception as e:
            print(f"  WARN: Google News fetch 실패 [{q_cfg['query']}]: {e}", file=sys.stderr)
            continue

        for entry in entries:
            link = entry.get("link", "")
            if not link or link in seen_urls:
                continue

            title = entry.get("title", "")
            pub = entry.get("published", "")

            # 날짜 필터
            if pub and not is_within_hours(pub, hours):
                continue

            # source 추출 (Google News RSS: title 뒤에 " - 언론사" 패턴)
            source_name = ""
            author = ""
            if hasattr(entry, "source") and entry.source:
                source_name = getattr(entry.source, "title", "") or ""
            if not source_name:
                # title 파싱: "기사 제목 - 매체명"
                m = re.search(r" - ([^-]+)$", title)
                if m:
                    source_name = m.group(1).strip()

            # author 필드 (RSS에 있으면)
            author = entry.get("author", "") or ""

            # 날짜 정규화
            pub_norm = ""
            if pub:
                try:
                    import email.utils
                    t = email.utils.parsedate_to_datetime(pub)
                    pub_norm = t.strftime("%Y-%m-%d")
                except Exception:
                    pub_norm = pub[:10] if len(pub) >= 10 else pub

            # 제목 또는 RSS snippet에 자사 키워드가 없는 기사는 skip.
            # Google News가 관련 매체의 다른 기사를 cluster로 묶어 반환할 때 노이즈가 섞임.
            combined = (title + " " + (entry.get("summary", "") or "")).lower()
            if not any(kw.lower() in combined for kw in SELF_KW):
                continue

            seen_urls.add(link)
            results.append({
                "url": link,
                "title": title,
                "source_name": source_name,
                "author": author,
                "published_date": pub_norm or TODAY,
                "full_text": "",
                "rss_snippet": entry.get("summary", "") or "",
                "_from_gnews": True,
                "is_domestic": q_cfg.get("domestic", True),
            })

        time.sleep(1.0)  # Google 레이트리밋 대응

    return results


# ── classify.py pr-articles 로드 ─────────────────────────────────

def load_classified_pr() -> list[dict]:
    """classify.py 가 생성한 pr-articles JSON 로드."""
    pr_path = PROCESSED_DIR / f"pr-articles-{TODAY}.json"
    if not pr_path.exists():
        return []
    raw = load_json(pr_path)
    if isinstance(raw, dict) and raw.get("articles"):
        return raw["articles"]
    if isinstance(raw, list):
        return raw
    return []


# ── GNews 기사 본문 fetch (full_text 없는 것만) ──────────────────

def fetch_missing_bodies(no_body: list[dict]) -> None:
    """본문 없는 GNews 기사의 full_text/author 를 채운다 (in-place).

    - 본문 캐시 (data/cache/body-cache.json) 우선 — rerun 시 네트워크 재fetch 방지
    - 캐시 미스만 병렬 fetch (worker 3 — gnews decode 레이트리밋 고려)
    """
    try:
        import trafilatura
        from lib.gnews_resolver import is_gnews_redirect
        from googlenewsdecoder import new_decoderv1
    except ImportError as _e:
        print(f"  WARN: 본문 fetch 스킵 ({_e})", file=sys.stderr)
        return

    import threading

    cache_path = CACHE_DIR / "body-cache.json"
    body_cache: dict = {}
    if cache_path.exists():
        try:
            body_cache = load_json(cache_path)
        except Exception:
            body_cache = {}

    # 캐시 TTL 30일 — ts 없는 구항목은 지금 스탬프 (다음 주기에 만료 대상)
    import time as _time
    _now = int(_time.time())
    _expired = [u for u, v in body_cache.items()
                if isinstance(v, dict) and v.get("ts", _now) < _now - 30 * 86400]
    for u in _expired:
        del body_cache[u]
    for v in body_cache.values():
        if isinstance(v, dict) and "ts" not in v:
            v["ts"] = _now
    if _expired:
        print(f"  본문 캐시 만료 정리: {len(_expired)}건 (30일 TTL)", file=sys.stderr)

    if not no_body:
        return
    print(f"  본문 fetch: {len(no_body)}건 (캐시 {len(body_cache)}건 보유)...", file=sys.stderr)

    # 1) 캐시 우선
    to_fetch: list[dict] = []
    cache_hit = 0
    for a in no_body:
        cached = body_cache.get(a["url"])
        if cached and cached.get("text"):
            a["full_text"] = cached["text"]
            if not a.get("author") and cached.get("author"):
                a["author"] = cached["author"]
            cache_hit += 1
        else:
            to_fetch.append(a)

    # 2) 캐시 미스만 네트워크 fetch (병렬)
    lock = threading.Lock()
    fetched_ok = 0
    cache_dirty = False

    def _fetch_one(a: dict) -> None:
        nonlocal fetched_ok, cache_dirty
        url = a["url"]
        try:
            fetch_url = url
            if is_gnews_redirect(fetch_url):
                result = new_decoderv1(fetch_url, interval=0.5)
                if result and result.get("status"):
                    fetch_url = result["decoded_url"]
                    # HTML에 노출되는 링크도 실제 기사 URL로 교체 (Google redirect 제거)
                    a["url"] = fetch_url
            downloaded = trafilatura.fetch_url(fetch_url)
            if downloaded:
                extracted_json = trafilatura.extract(
                    downloaded,
                    output_format="json",
                    with_metadata=True,
                    include_comments=False,
                    include_tables=False,
                )
                if extracted_json:
                    data = json.loads(extracted_json)
                    text = data.get("text") or data.get("raw_text") or ""
                    if text and len(text) > 100:
                        a["full_text"] = text
                        au = data.get("author") or ""
                        # trafilatura가 byline 떼간 경우 raw HTML서 복구
                        if "기자" not in au:
                            recovered = _author_from_raw_html(downloaded)
                            if recovered:
                                au = recovered
                        if not a.get("author") and au:
                            a["author"] = au
                        with lock:
                            fetched_ok += 1
                            body_cache[url] = {"text": text, "author": au, "ts": int(time.time())}
                            cache_dirty = True
        except Exception as e:
            print(f"  WARN: 기사 본문 fetch 실패 ({url[:60]}): {e}", file=sys.stderr)
        time.sleep(0.3)

    if to_fetch:
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=min(3, len(to_fetch))) as pool:
            list(pool.map(_fetch_one, to_fetch))

    print(f"  본문 fetch 완료: 신규 {fetched_ok}건 + 캐시 {cache_hit}건", file=sys.stderr)

    # 캐시 저장
    if cache_dirty:
        try:
            save_json(cache_path, body_cache)
        except Exception as _ce:
            print(f"  WARN: 본문 캐시 저장 실패: {_ce}", file=sys.stderr)


# ── 메인 ────────────────────────────────────────────────────────

def main():
    print(f"=== PR 모니터링 생성 ({TODAY}, {HOURS}h) ===", file=sys.stderr)

    # A. Google News RSS 직접 수집
    gnews_articles = fetch_gnews_pr(HOURS)

    # B. classify.py pr-articles 보완
    classified_pr = load_classified_pr()

    # 병합 (URL 기준 중복 제거, GNews 우선)
    seen: set[str] = set()
    merged: list[dict] = []

    for a in gnews_articles:
        url = a["url"]
        if url not in seen:
            seen.add(url)
            merged.append(a)

    for a in classified_pr:
        url = a.get("url", "")
        # 공유 캐시가 더 넓은 수집창(예: 뉴스레터 48h)일 수 있어 PR 창으로 재필터
        if not is_within_hours(a.get("published_date", ""), HOURS):
            continue
        if url and url not in seen:
            seen.add(url)
            merged.append(a)

    print(f"  합계: GNews {len(gnews_articles)}건 + classified {len(classified_pr)}건 "
          f"→ dedup 후 {len(merged)}건", file=sys.stderr)

    # C. GNews 기사 본문 fetch (full_text 없는 것만)
    fetch_missing_bodies(
        [a for a in merged if not a.get("full_text") and a.get("_from_gnews")]
    )

    # ── 개인 블로그 제외 (언론 매체 아님) ──────────────────────
    BLOG_DOMAINS = ("tistory.com", "blog.naver.com", "brunch.co.kr", "velog.io", "medium.com")
    before_b = len(merged)
    merged = [a for a in merged if not any(d in a.get("url", "") for d in BLOG_DOMAINS)]
    dropped_blog = before_b - len(merged)
    if dropped_blog:
        print(f"  블로그 제외: {dropped_blog}건 (티스토리 등 비매체)", file=sys.stderr)

    # ── 자사 무관 노이즈 제거 ──────────────────────────────────
    # 제목·본문 어디에도 자사명 없는 기사 제거 (단, 본문 미추출은 보존: 확인 불가)
    def _mentions_self(a: dict) -> bool:
        blob = (a.get("title", "") + " " + (a.get("full_text") or "")).lower()
        return any(kw in blob for kw in SELF_KW)
    before_n = len(merged)
    merged = [
        a for a in merged
        if _mentions_self(a) or not (a.get("full_text") or "")  # 본문 있는데 미언급 → 제거
    ]
    dropped_noise = before_n - len(merged)
    if dropped_noise:
        print(f"  노이즈 제거: 자사 무관 {dropped_noise}건 제외", file=sys.stderr)

    # ── 중복 제거 (제목 정규화 기준) ───────────────────────────
    def _norm_title(t: str) -> str:
        # 매체 접미사 ' - 매체명' 제거 + 공백/기호 정리 + 소문자
        t = re.sub(r"\s*-\s*[^-]+$", "", t)            # 끝의 ' - 출처' 제거
        t = re.sub(r"[\[\]\(\)“”\"'‘’·\s]", "", t)      # 기호·공백 제거
        return t.lower()[:40]
    seen_titles: dict[str, dict] = {}
    deduped: list[dict] = []
    for a in merged:
        key = _norm_title(a.get("title", ""))
        if not key:
            deduped.append(a)
            continue
        if key in seen_titles:
            # 중복: 본문 있는 쪽 우선 보존
            prev = seen_titles[key]
            if (a.get("full_text") or "") and not (prev.get("full_text") or ""):
                deduped[deduped.index(prev)] = a
                seen_titles[key] = a
            continue
        seen_titles[key] = a
        deduped.append(a)
    dropped_dup = len(merged) - len(deduped)
    if dropped_dup:
        print(f"  중복 제거: 유사 제목 {dropped_dup}건 통합", file=sys.stderr)
    merged = deduped

    # 각 기사 처리
    direct_rows: list[dict] = []    # 제목에 자사명 직접 포함
    indirect_rows: list[dict] = []  # 본문/맥락에만 간접 언급
    stock_rows: list[dict] = []

    for a in merged:
        title = a.get("title", "")
        body = a.get("full_text", "") or ""
        source = a.get("source_name", "") or a.get("_source_name", "")
        # 기자명 해석:
        #   1) 메타 author에 '기자' 있으면 신뢰(실제 byline) → 정규화
        #   2) 아니면(빈값·매체명) 본문 앞+뒤에서 '이름 기자' 추출
        #      (한국 기사는 byline이 본문 끝 이메일 옆에도 자주 위치)
        #   3) 둘 다 없으면 빈값 (매체명을 기자로 둔갑시키지 않음)
        _raw_author = (a.get("author", "") or a.get("_author", "") or "").strip()
        author = normalize_author(_raw_author) if "기자" in _raw_author else ""
        if not author and body:
            zone = body[:200] + "\n" + body[-300:]
            m = re.search(r"([가-힣]{2,4})\s*(?:기자|특파원)(?![가-힣])", zone)
            if m and m.group(1) not in MEDIA_NAMES:
                author = m.group(1) + " 기자"
        pub_date = a.get("published_date", TODAY)
        if pub_date and len(pub_date) > 10:
            pub_date = pub_date[:10]

        # source_name 도메인 정리
        url_str = a.get("url", "")
        if source:
            # "로봇신문 - 로봇신문" 같은 중복 제거
            source = re.sub(r"^(.+) - \1$", r"\1", source).strip()
        if not source or re.match(r"^[\w.-]+\.\w{2,}$", source):
            # domain 형태면 매핑 시도
            for domain, name in DOMAIN_TO_NAME.items():
                if domain in url_str:
                    source = name
                    break

        is_domestic = a.get("is_domestic", True)  # classified는 국내 기본
        # 국문/영문: 매체 접미사(' - 매일경제') 제외한 본제목 기준 (한글 출처명 오염 방지)
        _title_core = re.sub(r"\s*-\s*[^-]+$", "", title)
        is_korean = bool(re.search(r"[가-힣]", _title_core))

        tone = detect_tone(title, body)
        evidence = extract_evidence(title, body)
        summary = extract_summary(body, evidence)
        title_lower = title.lower()
        stock_in_title = any(k in title_lower for k in STOCK_KW)
        is_direct = any(kw in title_lower for kw in SELF_KW) and not stock_in_title
        stock = stock_in_title  # 제목 기준만 (본문 과민반응 방지)

        row = {
            "date": pub_date,
            "title": title,
            "source": source,
            "author": author,
            "url": url_str,
            "tone": tone,
            "_rule_tone": tone,  # 규칙 확정값 보존 — LLM 덮어쓰기 방지용
            "evidence": evidence,
            "summary": summary,
            "is_stock": stock,
            "is_direct": is_direct,
            "is_domestic": is_domestic,
            "is_korean": is_korean,
            "_body_head": body[:300],  # 직접 언급 요약(Haiku) 입력용
        }

        if stock:
            stock_rows.append(row)
        elif is_direct:
            direct_rows.append(row)
        else:
            indirect_rows.append(row)

    # 날짜 내림차순 정렬
    for lst in (direct_rows, indirect_rows, stock_rows):
        lst.sort(key=lambda x: x["date"], reverse=True)

    # 주가 섹션: 제목에 자사명 직접 포함된 것만 유지.
    # 자사 간접언급(제목에 자사명 없는 일반 시황·관련주)은 PR 가치 낮아 전부 제외.
    stock_direct = [r for r in stock_rows if any(kw in r["title"].lower() for kw in SELF_KW)]
    stock_indirect = []  # 제외
    dropped_stock = len(stock_rows) - len(stock_direct)
    if dropped_stock:
        print(f"  주가 노이즈 제거: 자사 비(非)직접언급 시황 {dropped_stock}건 제외", file=sys.stderr)
    stock_direct.sort(key=lambda x: x["date"], reverse=True)
    stock_rows = stock_direct

    all_rows = direct_rows + indirect_rows + stock_rows

    # ── 톤 판정(전체) + 직접 요약 — haiku 1회 통합 호출, 실패 시 규칙 유지 ──
    # all_rows = direct_rows + indirect_rows + stock_rows 이므로 앞 len(direct_rows)건이 요약 대상
    print("  톤 판정 + 직접 요약 중 (haiku 1회)...", file=sys.stderr)
    _tones, _summaries = classify_and_summarize_batch(all_rows, len(direct_rows))
    if _tones:
        for i, r in enumerate(all_rows):
            if i in _tones:
                # 규칙 기반이 이미 긍정/부정을 확정한 경우 LLM 덮어쓰기 방지.
                rule_tone = r.get("_rule_tone", "중립")
                if rule_tone in ("긍정", "부정"):
                    pass  # 규칙 확정 → LLM 무시
                else:
                    r["tone"] = _tones[i]
    for i, r in enumerate(direct_rows):
        r["llm_summary"] = _summaries.get(i) or (r.get("summary", "") or "")[:80]

    # ── CSV ────────────────────────────────────────────────────
    csv_path = PR_OUTPUT_DIR / f"pr-monitoring-{TODAY}.csv"
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=["날짜", "매체", "기자", "제목", "언급유형", "톤", "주가관련", "맥락", "URL"]
        )
        writer.writeheader()
        for r in all_rows:
            if r["is_stock"]:
                mention_type = "주가/시황"
            elif r.get("is_direct"):
                mention_type = "직접"
            else:
                mention_type = "간접"
            writer.writerow({
                "날짜": r["date"],
                "매체": r["source"],
                "기자": r["author"].replace(" 기자", ""),
                "제목": r["title"],
                "언급유형": mention_type,
                "톤": r["tone"],
                "주가관련": "Y" if r["is_stock"] else "",
                "맥락": r.get("evidence", ""),
                "URL": r["url"],
            })

    # ── XLSX ───────────────────────────────────────────────────
    xlsx_path = PR_OUTPUT_DIR / f"pr-monitoring-{TODAY}.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 기본 시트 제거

        COLS = ["날짜", "언급유형", "톤", "언어", "매체", "기자", "제목", "맥락", "URL"]
        HDR_FILL = PatternFill("solid", fgColor="292524")
        HDR_FONT = Font(bold=True, color="FAFAF9", size=10)
        THIN = Border(
            bottom=Side(style="thin", color="E7E5E4"),
            right=Side(style="thin", color="E7E5E4"),
        )
        TONE_COLORS = {"긍정": "DCFCE7", "부정": "FEE2E2", "중립": "F5F5F4"}

        def _mention_type(r: dict) -> str:
            if r["is_stock"]: return "주가/시황"
            if r.get("is_direct"): return "직접"
            return "간접"

        def _write_sheet(ws, rows: list[dict]):
            ws.append(COLS)
            for cell in ws[1]:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = Alignment(horizontal="center")
            for r in rows:
                mt = _mention_type(r)
                ws.append([
                    r["date"], mt, r["tone"],
                    "국문" if r.get("is_korean", True) else "영문",
                    r["source"], r["author"].replace(" 기자", ""), r["title"],
                    r.get("evidence", ""), r["url"],
                ])
                row_idx = ws.max_row
                tone_color = TONE_COLORS.get(r["tone"], "F5F5F4")
                for cell in ws[row_idx]:
                    cell.border = THIN
                    if cell.column == 3:  # 톤 컬럼 배경
                        cell.fill = PatternFill("solid", fgColor=tone_color)
            # 컬럼 너비
            widths = [12, 10, 8, 6, 18, 12, 50, 50, 40]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

        tabs = [
            ("간접 언급", indirect_rows),
            ("직접 언급", direct_rows),
            ("주가·시황", stock_rows),
            ("전체", all_rows),
        ]
        for tab_name, rows in tabs:
            if not rows and tab_name != "전체":
                continue
            ws = wb.create_sheet(tab_name)
            _write_sheet(ws, rows)

        wb.save(xlsx_path)
        print(f"  xlsx: {xlsx_path.name} ({len(all_rows)}행)", file=sys.stderr)
    except Exception as e:
        print(f"  WARN: xlsx 생성 실패: {e}", file=sys.stderr)
        xlsx_path = None

    # ── HTML ───────────────────────────────────────────────────
    TONE_STYLES = {
        "부정": ("#fef2f2", "#dc2626", "#fecaca"),
        "긍정": ("#f0fdf4", "#16a34a", "#bbf7d0"),
        "중립": ("#f5f5f4", "#78716c", "#e7e5e4"),
    }

    def render_article(row: dict) -> str:
        bg, fg, border = TONE_STYLES.get(row["tone"], TONE_STYLES["중립"])
        tone_badge = (
            f'<span style="display:inline-block;font-size:11px;font-weight:600;'
            f'padding:1px 8px;border-radius:10px;background:{bg};color:{fg};'
            f'border:1px solid {border};">{row["tone"]}</span>'
        )
        geo_badge = (
            '<span style="font-size:11px;color:#9ca3af;margin-right:5px;">국문</span>'
            if row.get("is_korean", True) else
            '<span style="font-size:11px;color:#6b7280;font-weight:600;margin-right:5px;">영문</span>'
        )
        source_line = geo_badge + row["source"]
        if row.get("author"):
            source_line += f' · {row["author"]}'
        source_line += f' · {row["date"]} · {tone_badge}'

        ev = row.get("evidence", "")
        summ = row.get("llm_summary", "") or ""
        is_stock = row.get("is_stock")
        is_direct = row.get("is_direct")
        is_indirect = not is_stock and not is_direct

        def _annot_block(label: str, content: str) -> str:
            return (
                f'<blockquote style="margin:4px 0 0;padding:5px 12px;border-left:3px solid #d6d3d1;'
                f'font-size:12px;color:#57534e;font-style:italic;">'
                f'<span style="font-size:10px;font-weight:600;color:#78716c;'
                f'text-transform:uppercase;letter-spacing:1px;margin-right:6px;">{label}</span>'
                f'{content}</blockquote>\n'
            )

        # 직접 언급: 자사가 기사 주체 → 언급위치 무의미. 1줄 요약(Haiku)을 보여준다.
        # 간접 언급·주가/시황: 자사가 어디서 언급됐는지(언급 위치)가 유용.
        if is_direct:
            ev_html = _annot_block("요약", summ) if summ else ""
        elif ev:
            ev_html = _annot_block("언급 위치", ev)
        elif is_indirect:
            ev_html = (
                '<div style="margin:6px 0 0;font-size:11px;color:#a8a29e;'
                'font-style:italic;">본문 미추출 · 원문에서 언급 위치 확인 필요</div>\n'
            )
        else:
            ev_html = ""

        return (
            f'<div style="border:1px solid #e7e5e4;border-radius:8px;'
            f'padding:14px 18px;margin-bottom:10px;">\n'
            f'  <div style="font-size:14px;font-weight:700;margin-bottom:2px;">'
            f'<a href="{row["url"]}" style="color:#2563eb;text-decoration:none;">'
            f'{row["title"]}</a></div>\n'
            f'  <div style="font-size:12px;color:#78716c;margin-bottom:4px;">'
            f'{source_line}</div>\n'
            f'{ev_html}'
            f'</div>\n'
        )

    def render_section(title: str, rows: list, empty_msg: str = "") -> str:
        if not rows and not empty_msg:
            return ""
        if not rows:
            items = f'<p style="color:#78716c;font-size:13px;">{empty_msg}</p>'
            return (
                f'<h2 style="font-size:16px;font-weight:700;color:#44403c;margin:0 0 12px;">'
                f'{title}</h2>\n{items}'
            )
        # 국문 먼저, 영문 뒤 (소제목 없이 연속 배치)
        korean = [r for r in rows if r.get("is_korean", True)]
        english = [r for r in rows if not r.get("is_korean", True)]
        parts = "".join(render_article(r) for r in (korean + english))
        return (
            f'<h2 style="font-size:16px;font-weight:700;color:#44403c;margin:0 0 12px;">'
            f'{title}</h2>\n{parts}'
        )

    stock_direct_count = len(stock_direct)
    direct_note = ""
    if stock_direct_count:
        direct_note = (
            f'<div style="font-size:11px;color:#a8a29e;margin-bottom:16px;font-style:italic;">'
            f'※ 제목에 자사명이 포함된 기사 중 주가·시황 성격의 {stock_direct_count}건은 편집 기사로 분류하지 않고 하단의 별도 섹션에 표시됩니다.</div>\n'
        )
    direct_section = direct_note + render_section(
        f"직접 언급 ({len(direct_rows)}건)",
        direct_rows,
        "직접 언급 기사 없음 (주가 관련 직접언급은 아래 주가·시황 섹션 참고)",
    )
    indirect_section = ""
    if indirect_rows:
        indirect_section = (
            '<hr style="border:none;border-top:1px solid #e7e5e4;margin:30px 0 20px;">\n'
            + render_section(f"간접 언급 ({len(indirect_rows)}건)", indirect_rows)
        )
    stock_section_html = ""
    if stock_rows:
        body = "".join(render_article(r) for r in stock_direct)
        stock_section_html = (
            '<hr style="border:none;border-top:1px solid #e7e5e4;margin:30px 0 20px;">\n'
            '<h2 style="font-size:16px;font-weight:700;color:#78716c;margin:0 0 12px;">'
            f'주가 · 시황 관련 ({len(stock_rows)}건)<span style="font-size:11px;'
            f'font-weight:400;color:#a8a29e;margin-left:8px;">제목에 자사명 포함된 기사만</span></h2>\n'
            f'{body}'
        )

    # ── 전체 현황 브리핑 (LLM 2-3줄 요약) ──────────────────────
    def build_top_narrative() -> str:
        ed_rows = direct_rows + indirect_rows
        total = len(ed_rows) + len(stock_rows)
        if not ed_rows and not stock_rows:
            return ""

        # LLM 프롬프트: 편집 기사 제목+언급위치 + 주가 주요 제목
        prompt_lines = [
            f"당신은 {_BRANDING['org_name']} 마케팅팀 PR 담당자입니다.",
            f"오늘({TODAY}) 자사 보도 모니터링 결과를 바탕으로 '오늘의 PR 동향 브리핑'을 작성하세요.",
            "",
            "작성 규칙:",
            "- 정확히 3개 문단. 문단 사이는 빈 줄로 구분. (4개 이상 금지)",
            "- 각 문단 1~3문장. 여러 기사를 '종합'해 큰 흐름을 말하라. 기사별로 나열하지 마라.",
            "- **한 문단 = 한 주제.** 무관한 사건을 같은 문단에 잇지 마라. 같은 사건의 전개는 한 문단에 모아라 (수사 얘기가 1·3문단에 쪼개지면 안 된다).",
            "- 권장 구성: ①오늘 지배 이슈 — 가장 큰 사건의 전말을 한 문단에 완결 ②그 외 자사 보도 — 별건들. 비중 낮으면 '이와 별도로 ~' 한 문장씩 ③주가·종합 한 문장.",
            "- 핵심 사건·동반 기업·이벤트명은 언급해도 좋다. 단, 어느 매체가 보도했는지(매체명·출처 도메인)는 절대 나열하지 마라.",
            "- '비즈니스플러스·매일경제가 보도', 'jabon.co.kr에 따르면' 같은 출처 열거 금지. 보도 주체가 아니라 '무슨 일이 있었는지'를 써라.",
            "- 지엽적 디테일(개별 수치 나열, 마이너 사실)보다 'PR 관점의 큰 그림'을 우선.",
            "- 주가·시황은 비중 낮음: 별도 문단 만들지 말고 필요시 한 문장만. 등락 수치 나열 금지.",
            "- 담당자가 이 브리핑만 읽고 '오늘 우리 이름이 어떤 맥락으로 돌았는지' 큰 그림을 잡을 수 있어야 함.",
            "",
            "비약 금지 (위반 시 브리핑 무효):",
            "- 기사에 적힌 사실만 쓴다. 기사가 직접 말하지 않은 결론·전망·평가를 만들지 마라.",
            "- 금지 단어: 프레임 / 흐름 / 신호 / 시사 / 모멘텀 / 존재감 / 기회 요인 / 리스크 요인 / 연동 / 가시화 / 레퍼런스 / 내러티브 / 관전 포인트.",
            "- 금지 표현: '~을 시사한다', '~이 필요한 시점', '~을 확인하는', '~로 해석된다', '~할 전망', '~이 기대된다'.",
            "- 해석이 필요하면 '~라는 뜻이다' 형태로 기사 사실에 직결시켜라. 사실과 안 이어지면 쓰지 마라.",
            "- 한 문장 60자 이내. 복문은 쪼개라.",
            "- 예: ✅ '세이프틱스 상대 무효심판 2건이 인용 확정됐다. 특허 분쟁이 종결됐다는 뜻이다.'",
            "- 예: ❌ '법적 리스크가 해소되며 상용화 레퍼런스도 함께 쌓이는 흐름이다.'",
            "- 한국어. 브리핑 텍스트만 출력. 제목·서론·부연설명 없이. 문단 사이는 빈 줄로 구분.",
            "",
            f"[수집 범위: 최근 {HOURS}h / 편집 기사 {len(ed_rows)}건 / 주가·시황 {len(stock_rows)}건]",
            "---",
            "## 편집 기사 (자사 이름 언급됨)",
        ]

        # 주제 다양성 보장: 동일 키워드 묶음은 최대 3건으로 cap 후 나머지 슬롯을 다른 주제로 채움
        TOPIC_KEYS = [
            (list(tk["keywords"]), tk["label"]) for tk in _PR_QUERIES["topic_keys"]
        ]

        def _topic(title: str) -> str:
            t = title.lower()
            for kws, label in TOPIC_KEYS:
                if any(k in t for k in kws):
                    return label
            return "기타"

        topic_counts: dict = {}
        diverse_rows: list = []
        deferred: list = []
        for r in ed_rows:
            tp = _topic(r.get("title", ""))
            if topic_counts.get(tp, 0) < 3:
                topic_counts[tp] = topic_counts.get(tp, 0) + 1
                diverse_rows.append(r)
            else:
                deferred.append(r)
        # 20개 미만이면 deferred에서 보충 (주제 cap 초과분)
        for r in deferred:
            if len(diverse_rows) >= 20:
                break
            diverse_rows.append(r)

        for r in diverse_rows[:20]:
            ev = r.get("evidence", "")
            tone = r.get("tone", "")
            prompt_lines.append(f"- [{tone}] {r['title']}" + (f"\n  언급: {ev[:120]}" if ev else ""))
        if stock_rows:
            stock_direct = [r for r in stock_rows if any(kw in r["title"].lower() for kw in SELF_KW)]
            prompt_lines.append("\n## 주가·시황 (자사 직접 언급)")
            for r in stock_direct[:8]:
                prompt_lines.append(f"- {r['title']}")
            if not stock_direct:
                prompt_lines.append(f"- 자사 직접 언급 없음 (시황 기사 {len(stock_rows)}건)")

        narrative = ""
        try:
            if not CLAUDE_BIN:
                raise RuntimeError("claude CLI 미설치 — 규칙 fallback")
            # 90초는 CLI 콜드스타트+생성에 부족해 타임아웃 fallback(통계 한 줄)이
            # 자주 발생했음 → 300초. 핵심 산출물이므로 1회 재시도.
            for attempt in (1, 2):
                try:
                    res = subprocess.run(
                        [CLAUDE_BIN, "-p", "\n".join(prompt_lines), "--model", "sonnet"],
                        capture_output=True, text=True, timeout=300,
                    )
                    if res.returncode == 0 and res.stdout.strip():
                        narrative = res.stdout.strip()
                        break
                    print(f"  WARN: 상단 브리핑 시도 {attempt} 실패 (rc={res.returncode})",
                          file=sys.stderr)
                except subprocess.TimeoutExpired:
                    print(f"  WARN: 상단 브리핑 시도 {attempt} 타임아웃(300s)", file=sys.stderr)
        except Exception as e:
            print(f"  WARN: 상단 브리핑 생성 실패: {e}", file=sys.stderr)

        if not narrative:
            # fallback: 간단 통계 텍스트
            narrative = f"최근 {HOURS}h 기준 총 {total}건 수집. 편집 언급 {len(ed_rows)}건, 주가·시황 {len(stock_rows)}건."

        # 문단(빈 줄 기준) 분리 → 각 문단을 개별 줄로 렌더
        paras = [p.strip() for p in re.split(r"\n\s*\n", narrative) if p.strip()]
        if len(paras) <= 1:
            # 빈 줄 없으면 단일 줄바꿈으로도 시도
            paras = [p.strip() for p in narrative.split("\n") if p.strip()]
        body_html = "".join(
            f'<div style="font-size:13px;color:#44403c;line-height:1.6;'
            f'margin-bottom:{"8px" if i < len(paras)-1 else "0"};">{p}</div>\n'
            for i, p in enumerate(paras)
        )

        return (
            '<div style="background:#fafaf9;border:1px solid #e7e5e4;border-radius:8px;'
            'padding:14px 18px;margin-bottom:24px;">\n'
            '<div style="font-size:11px;font-weight:700;color:#78716c;'
            'letter-spacing:1px;margin-bottom:8px;">오늘의 브리핑</div>\n'
            f'{body_html}'
            '</div>\n'
        )

    headline_digest = build_top_narrative()

    # ── 요약 테이블 ─────────────────────────────────────────────
    def tone_counts(rows: list) -> dict:
        c = {"긍정": 0, "중립": 0, "부정": 0}
        for r in rows:
            c[r.get("tone", "중립")] = c.get(r.get("tone", "중립"), 0) + 1
        return c

    def tone_badges(counts: dict) -> str:
        parts = []
        if counts["긍정"]:
            parts.append(f'<span style="color:#16a34a;font-weight:600;">긍정 {counts["긍정"]}</span>')
        if counts["부정"]:
            parts.append(f'<span style="color:#dc2626;font-weight:600;">부정 {counts["부정"]}</span>')
        if counts["중립"]:
            parts.append(f'<span style="color:#78716c;">중립 {counts["중립"]}</span>')
        return " · ".join(parts) if parts else "—"

    cats = [
        ("직접 언급", direct_rows),
        ("간접 언급", indirect_rows),
        ("주가 · 시황", stock_rows),
    ]
    total_all = len(direct_rows) + len(indirect_rows) + len(stock_rows)
    total_counts = tone_counts(all_rows)
    korean_total = sum(1 for r in all_rows if r.get("is_korean", True))
    english_total = total_all - korean_total

    table_rows_html = ""
    for label, rows in cats:
        if not rows:
            continue
        tc = tone_counts(rows)
        table_rows_html += (
            f'<tr>'
            f'<td style="padding:6px 12px;font-weight:600;border-bottom:1px solid #f5f5f4;">{label}</td>'
            f'<td style="padding:6px 12px;text-align:center;border-bottom:1px solid #f5f5f4;">{len(rows)}</td>'
            f'<td style="padding:6px 12px;border-bottom:1px solid #f5f5f4;font-size:12px;">{tone_badges(tc)}</td>'
            f'</tr>\n'
        )
    table_rows_html += (
        f'<tr style="background:#f5f5f4;">'
        f'<td style="padding:6px 12px;font-weight:700;">계</td>'
        f'<td style="padding:6px 12px;text-align:center;font-weight:700;">{total_all}</td>'
        f'<td style="padding:6px 12px;font-size:12px;">{tone_badges(total_counts)}</td>'
        f'</tr>\n'
    )

    summary_table = (
        '<div style="background:#fafaf9;border:1px solid #e7e5e4;border-radius:8px;'
        'padding:16px 18px;margin-bottom:28px;">\n'
        '<div style="font-size:12px;font-weight:700;color:#78716c;'
        'letter-spacing:1px;margin-bottom:10px;">모니터링 요약</div>\n'
        '<table style="width:100%;border-collapse:collapse;font-size:13px;">\n'
        '<thead><tr style="border-bottom:2px solid #e7e5e4;">'
        '<th style="padding:4px 12px;text-align:left;font-size:11px;color:#78716c;font-weight:600;">카테고리</th>'
        '<th style="padding:4px 12px;text-align:center;font-size:11px;color:#78716c;font-weight:600;">건수</th>'
        '<th style="padding:4px 12px;text-align:left;font-size:11px;color:#78716c;font-weight:600;">톤</th>'
        '</tr></thead>\n'
        f'<tbody>{table_rows_html}</tbody>\n'
        '</table>\n'
        f'<div style="font-size:11px;color:#b7b3ae;margin-top:8px;">국문 {korean_total}건 · 영문 {english_total}건</div>\n'
        '</div>\n'
    )

    total_editorial = len(direct_rows) + len(indirect_rows)

    last_run_file = PR_OUTPUT_DIR / ".last-pr-monitor-run"
    prev_run = last_run_file.read_text(encoding="utf-8").strip() if last_run_file.exists() else ""
    generated = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d %H:%M")
    prev_run_html = (
        f'<div style="margin-top:2px;">직전 발행: {prev_run}</div>'
        if prev_run else ""
    )

    html = (
        '<!DOCTYPE html>\n<html lang="ko"><head><meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
        f'<title>PR 모니터링 — {TODAY}</title>\n'
        '<style>@import url("https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;800&display=swap");\n'
        'body { font-family: "Noto Sans KR", -apple-system, "Pretendard", sans-serif; '
        'max-width: 700px; margin: 0 auto; padding: 30px 20px; '
        'color: #1c1917; line-height: 1.7; background: #fff; }</style>'
        '</head><body>\n'
        '<div style="font-size:11px;font-weight:700;color:#78716c;'
        f'letter-spacing:3px;margin-bottom:4px;">{_BRANDING["html_header_pr"]}</div>\n'
        '<h1 style="font-size:22px;font-weight:800;margin:0 0 4px;">자사 보도 모니터링</h1>\n'
        f'<div style="font-size:13px;color:#78716c;margin-bottom:20px;">'
        f'{TODAY} · 최근 {HOURS}h</div>\n'
        f'{headline_digest}'
        f'{summary_table}'
        f'{direct_section}'
        f'{indirect_section}'
        f'{stock_section_html}'
        '<div style="margin-top:30px;padding-top:12px;border-top:1px solid #e7e5e4;'
        'font-size:11px;color:#a8a29e;">'
        f'{_BRANDING["html_footer"]}'
        f'<div style="margin-top:3px;">자동생성 {generated} KST</div>'
        f'{prev_run_html}'
        '</div>\n'
        '</body></html>'
    )

    out_html = PR_OUTPUT_DIR / f"pr-monitoring-{TODAY}.html"
    out_html.parent.mkdir(parents=True, exist_ok=True)
    out_html.write_text(html, encoding="utf-8")
    last_run_file.write_text(generated, encoding="utf-8")

    print(f"✅ PR Monitor: 직접 {len(direct_rows)}건 · 간접 {len(indirect_rows)}건 · 주가 {len(stock_rows)}건 → {out_html}")


if __name__ == "__main__":
    main()
